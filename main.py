#Importando bibliotecas
import pyautogui
import pyperclip
import time
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import shutil
import os
import sys

#Importando utilitários
from utils.connection import checar_conectividade
from utils.input import input_numres
from utils.validate import validate_status, validate_resolution, validate_excel, validate_folders, validate_download, validate_input
from utils.filesystem import create_directory, copy_excel_file

#Importando configurações
from config.settings import DOCUMENTOS_DIR, CONTROLE_EXCEL, DOWNLOADS_DIR

status = checar_conectividade()

validate_status(status)

validate_resolution()

TIMEOUT = 60

files_directory = DOWNLOADS_DIR

#Verificando pasta downloads
validate_folders(files_directory)

validate_download(files_directory)

pyautogui.FAILSAFE = True

dia, mes, ano = validade_date()

validate_excel()

#Entrada do número de registros
while True:
    try:
        num = input_numres()
        num_validado = validate_input(num)
        break
    except ValueError as e:
        print(f"Erro: {e}")

#minimizar vscode
validate_vscode()

#POSIÇÕES
#POSICAO INICIAL LINHA SIGAMA = 336

PNx, PNy = pyautogui.locateCenterOnScreen('./image/nome_image.png', confidence= 0.6)
posicao_i_nome_S = [PNx, PNy + 40]
posicao_a_nome_S = posicao_i_nome_S

# posicao_i_nome_E = [252, 309]
# posicao_a_nome_E = posicao_i_nome_E

PCx, PCy = pyautogui.locateCenterOnScreen('./image/cpf_image.png', confidence= 0.6)
posicao_i_cpf_S = [PCx, PCy + 40]
posicao_a_cpf_S = posicao_i_cpf_S

# posicao_i_cpf_E = [590, 309]
# posicao_a_cpf_E = posicao_i_cpf_E

PLx, PLy = pyautogui.locateCenterOnScreen('./image/operacoes_image.png', confidence= 0.5)
posicao_i_lupa = [PLx + 52, PLy + 34]
posicao_a_lupa = posicao_i_lupa

#Verificando pasta documentos
validate_folders(DOCUMENTOS_DIR)

#Criar e manipular pastas
diretorio = create_directory(DOCUMENTOS_DIR, ano, mes, dia)
copy_excel_file(CONTROLE_EXCEL, diretorio)


for j in range(num):

    status = checar_conectividade()

    validate_status(status)

    #copiar nome - SIGAMA
    pyautogui.moveTo(posicao_a_nome_S)
    pyautogui.tripleClick()
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.5)

    arquivo = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",str(ano), str(mes_nome), str(dia), "Controle de Solicitação.xlsx")
    coluna = 'A'  # coluna desejada
    nome = pyperclip.paste()

    wb = load_workbook(arquivo)
    ws = wb["Controle de Solicitação"] # ou wb["NomeDaAba"]

    linha = 1

    # encontra a primeira célula vazia
    while ws[f"{coluna}{linha}"].value is not None:
        linha += 1

    # escreve o nome
    ws[f"{coluna}{linha}"] = nome

    # salva
    wb.save(arquivo)

    #copiar cpf SIGAMA
    pyautogui.moveTo(posicao_a_cpf_S)
    pyautogui.doubleClick()
    pyautogui.hotkey('ctrl', 'c')
    cpf = pyperclip.paste()

    #cola cpf no excel
    coluna = 'B'

    # escreve o cpf
    ws[f"{coluna}{linha}"] = cpf

    # salva
    wb.save(arquivo)

    colunas = ['A', 'B']   # colunas de destino

    #criar pasta com nome e cpf
    nome_cpf = [ws[f"{col}{linha}"].value for col in colunas]
    nome_cpf_tratado = f"{nome} - {cpf}"

    Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso", str(ano), str(mes_nome), str(dia), nome_cpf_tratado).mkdir(exist_ok=True)

    #localizar lupa
    pyautogui.moveTo(posicao_a_lupa)
    pyautogui.click()
    
    status = checar_conectividade()

    if status != "OK":
        print(f"Problema de conexão: {status}")
        sys.exit()

    time.sleep(0.2)

    #Localizar anexo de documentos
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('./image/anexo_image.png', confidence = 0.8))
    x1, y1 = pyautogui.locateCenterOnScreen('./image/anexo_image.png', confidence = 0.8)
    x1 = x1 - 25

    #clicar nos documentos
    pyautogui.moveTo(x1, y1+27)
    for i in range(5):
        y1 = y1+23
        pyautogui.moveTo(x1, y1)
        time.sleep(0.5)
        pyautogui.click()
        time.sleep(0.2)
    time.sleep(0.2)
    pyautogui.click(pyautogui.locateCenterOnScreen('./image/X_image.png', confidence = 0.8))

    #Colocar documentos na pasta
    destino = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia),
    nome_cpf_tratado
)
    
    hoje = date.today()
    agora = datetime.now()

    #identificar erro no diretório de arquivos
    for arquivo in files_directory.iterdir():
        if arquivo.is_file():
            if arquivo.suffix != '.crdownload':
                data_arquivo = datetime.fromtimestamp(arquivo.stat().st_mtime)
                if agora - data_arquivo < timedelta(seconds=15):
                    time.sleep(0.2)
                    shutil.move(arquivo, destino / arquivo.name)
            else:
                print("Erro no download")
                sys.exit()

    #iteração das posições
    posicao_a_nome_S[1] = posicao_a_nome_S[1] + 40

    posicao_a_cpf_S[1] = posicao_a_cpf_S[1] + 40

    posicao_a_lupa[1] = posicao_a_lupa[1] + 40

#Abrir paginas no final da execução
caminho_pasta = Path(
    "Z:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia)
)

caminho_excel = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",str(ano), str(mes_nome), str(dia), "Controle de Solicitação.xlsx")

# os.startfile(caminho_csv)
os.startfile(caminho_pasta)
os.startfile(caminho_excel)