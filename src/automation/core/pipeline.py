#Importando bibliotecas
import pyautogui
import pyperclip
import time
from openpyxl import load_workbook
from pathlib import Path
import os

#Importando utilitários
from automation.utils.connection import checar_conectividade
from automation.utils.validate import validate_status, validate_resolution, validate_excel, validate_folders, validate_downloads_folder, validate_input, validate_vscode, validade_date, validate_position, validate_click
from automation.utils.filesystem import create_directory, copy_excel_file
from automation.utils.position import get_name_positions, get_cpf_positions, get_lupa_position, get_docs_position, get_documents
from automation.core.controller import RunController, StopRequested

#Importando configurações
from automation.config.settings import DOCUMENTOS_DIR, CONTROLE_EXCEL, DOWNLOADS_DIR, IMAGE_DIR

def run_pipeline(controller: RunController, num_registros: int):
    
        status = checar_conectividade()

        validate_status(status)

        validate_resolution()

        TIMEOUT = 60

        files_directory = DOWNLOADS_DIR

        images_directory = IMAGE_DIR

        #Verificando pasta downloads
        validate_folders(files_directory)

        #validate_download(files_directory)

        pyautogui.FAILSAFE = True

        dia, mes, ano = validade_date()

        validate_excel()

        controller.checkpoint("inicio_pipeline")

        position_name = get_name_positions(images_directory)

        position_cpf = get_cpf_positions(images_directory)

        position_lupa = get_lupa_position(images_directory)

        #Verificando pasta documentos
        validate_folders(DOCUMENTOS_DIR)

        #Criar e manipular pastas
        diretorio = create_directory(DOCUMENTOS_DIR, ano, mes, dia)
        copy_excel_file(CONTROLE_EXCEL, diretorio)

        for j in range(num_registros):

            controller.checkpoint(f"inicio_iteracao_{j}")

            status = checar_conectividade()

            validate_status(status)

            #copiar nome - SIGAMA

            arquivo = Path(DOCUMENTOS_DIR, str(ano), str(mes), str(dia), "Controle de Solicitação.xlsx")
            coluna = 'A'  # coluna desejada
            controller.checkpoint("antes_nome")
            nome = validate_position(position_name, str(3))

            wb = load_workbook(arquivo)
            ws = wb["Controle de Solicitação"] # ou wb["NomeDaAba"]

            linha = 1

            # encontra a primeira célula vazia
            while ws[f"{coluna}{linha}"].value is not None:
                linha += 1

            # escreve o nome
            ws[f"{coluna}{linha}"] = nome

            # salva
            # wb.save(arquivo) - Acredito que só necessite de uma chamada de save no final

            #copiar cpf SIGAMA
            controller.checkpoint("antes_cpf")
            cpf = validate_position(position_cpf, str(2))

            #inicializa a coluna cpf como B no excel
            coluna = 'B'

            # escreve o cpf
            ws[f"{coluna}{linha}"] = cpf

            # salva
            controller.checkpoint("antes_salvar_excel")
            wb.save(arquivo)

            colunas = ['A', 'B']   # colunas de destino

            #criar pasta com nome e cpf
            nome_cpf = [ws[f"{col}{linha}"].value for col in colunas]
            nome_cpf_tratado = f"{nome} - {cpf}"

            Path(DOCUMENTOS_DIR, str(ano), str(mes), str(dia), nome_cpf_tratado).mkdir(exist_ok=True)

            #localizar lupa
            controller.checkpoint("antes_lupa")
            validate_position(position_lupa, str(1))
            
            status = checar_conectividade()

            validate_status(status)

            time.sleep(0.2)

            #Localizar anexo de documentos
            controller.checkpoint("antes_docs_pos")
            x1, y1 = get_docs_position(images_directory)

            #clicar nos documentos
            controller.checkpoint("antes_baixar_docs")
            get_documents(x1, y1, files_directory, DOCUMENTOS_DIR, ano, mes, dia, nome_cpf_tratado, images_directory)

            #iteração das posições
            position_name[1] = position_name[1] + 40

            position_cpf[1] = position_cpf[1] + 40

            position_lupa[1] = position_lupa[1] + 40

        #Abrir paginas no final da execução
        caminho_pasta = Path(DOCUMENTOS_DIR, str(ano), str(mes), str(dia))

        caminho_excel = Path(DOCUMENTOS_DIR, str(ano), str(mes), str(dia), "Controle de Solicitação.xlsx")

        # os.startfile(caminho_csv)
        os.startfile(caminho_pasta)
        os.startfile(caminho_excel)
