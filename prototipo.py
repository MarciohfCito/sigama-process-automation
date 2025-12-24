import pyautogui
import pyperclip
import time
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import shutil
import os
import sys
import win32com.client as win32
import socket
import requests

def internet_ativa(timeout=3):
    try:
        socket.create_connection(("8.8.8.8", 53), timeout=timeout)
        return True
    except OSError:
        return False

def sigama_online():
    try:
        r = requests.get("https://sigama.aged.ma.gov.br/", timeout=5)
        return r.status_code < 500
    except:
        return False

import requests

def internet_http(url="https://www.google.com", timeout=5):
    try:
        r = requests.get(url, timeout=timeout)
        return r.status_code == 200
    except requests.RequestException:
        return False

def checar_conectividade():
    if not internet_ativa():
        return "SEM_INTERNET"
    if not internet_http():
        return "INTERNET_INSTAVEL"
    if not sigama_online():
        return "SIGAMA caiu"
    return "OK"

def fechar_excel_se_aberto():
    try:
        excel = win32.GetActiveObject("Excel.Application")
        print("📊 Excel aberto encontrado")

        for wb in excel.Workbooks:
            if not wb.Saved:
                wb.Save()
                print(f"💾 Salvo: {wb.Name}")

        excel.Quit()
        print("❌ Excel fechado com sucesso")

    except Exception:
        print("✅ Excel não está aberto")

status = checar_conectividade()

if status != "OK":
    print(f"⚠️ Problema de conexão: {status}")
    sys.exit()

TIMEOUT = 60  # segundos

downloads = Path("C:/Users/marcio.cito/Downloads")

def esperar_qualquer_download(timeout=60):
    inicio = time.time()
    arquivos_iniciais = set(os.listdir(downloads))

    while True:
        arquivos_atuais = set(os.listdir(downloads))
        novos = arquivos_atuais - arquivos_iniciais

        for arquivo in novos:
            if not arquivo.endswith('.crdownload'):
                return os.path.join(downloads, arquivo)
        
        if time.time() - inicio > timeout:
            raise TimeoutError('⏰ Nenhum download detectado')

pyautogui.FAILSAFE = True

data_atual = datetime.today()


dia = data_atual.strftime("%d")
mes = data_atual.month
ano = data_atual.strftime("%Y")

meses = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}

mes_nome = meses[mes]

print(dia, mes_nome, ano)

fechar_excel_se_aberto()

#inserir o numero de registros
num = int(input("Digite o número de registros: "))

def minimizar():
    pyautogui.moveTo(pyautogui.moveTo(x=1806, y=7))
    pyautogui.click()
    time.sleep(0.5)

#minimizar vscode
minimizar()
time.sleep(0.5)

#POSIÇÕES
#POSICAO INICIAL LINHA SIGAMA = 336

posicao_i_nome_S = [77, 332]
posicao_a_nome_S = posicao_i_nome_S

# posicao_i_nome_E = [252, 309]
# posicao_a_nome_E = posicao_i_nome_E

PCx, PCy = pyautogui.locateCenterOnScreen('./image/cpf_image.png', confidence= 0.6)
posicao_i_cpf_S = [PCx, PCy + 40]
posicao_a_cpf_S = posicao_i_cpf_S

# posicao_i_cpf_E = [590, 309]
# posicao_a_cpf_E = posicao_i_cpf_E

PLx, PLy = pyautogui.locateCenterOnScreen('./image/operacoes_image.png', confidence= 0.5)
posicao_i_lupa = [PLx + 52, PLy + 34]
posicao_a_lupa = posicao_i_lupa

#Criar pastas

base = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso")
pasta = base / ano / mes_nome / dia
pasta.mkdir(parents=True, exist_ok=True)

origem = Path("Z:/SIGAMA/Controle de Solicitação.xlsx")
destino_pasta = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",str(ano), str(mes_nome), str(dia))

destino_arquivo = destino_pasta / origem.name

if not destino_arquivo.exists():
    shutil.copy(origem, destino_pasta)
    print("Arquivo copiado com sucesso")
else:
    print("Arquivo já existe, não foi copiado")

for j in range(num):

    if not internet_ativa():
        print("❌ Internet indisponível")
        sys.exit()

    if not sigama_online():
        print("⚠️ SIGAMA fora do ar ou lento")
        sys.exit()
        # tenta novamente ou encerra

    #copiar nome - SIGAMA
    pyautogui.moveTo(posicao_a_nome_S)
    pyautogui.tripleClick()
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.5)

    # --------- USANDO EXCEL ---------------
    arquivo = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",str(ano), str(mes_nome), str(dia), "Controle de Solicitação.xlsx")
    coluna = 'A'  # coluna desejada
    nome = pyperclip.paste()

    wb = load_workbook(arquivo)
    ws = wb["Controle de Solicitação"] # ou wb["NomeDaAba"]

    linha = 1

    # encontra a primeira célula vazia
    while ws[f"{coluna}{linha}"].value is not None:
        linha += 1

    # escreve o nome
    ws[f"{coluna}{linha}"] = nome

    # salva
    wb.save(arquivo)

    #copiar cpf SIGAMA
    pyautogui.moveTo(posicao_a_cpf_S)
    pyautogui.doubleClick()
    pyautogui.hotkey('ctrl', 'c')
    cpf = pyperclip.paste()
    # #cola cpf no excel
    coluna = 'B'

    # escreve o cpf
    ws[f"{coluna}{linha}"] = cpf

    # salva
    wb.save(arquivo)

    colunas = ['A', 'B']   # colunas de destino

    #criar pasta com nome e cpf
    nome_cpf = [ws[f"{col}{linha}"].value for col in colunas]
    nome_cpf_tratado = f"{nome} - {cpf}"

    Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso", str(ano), str(mes_nome), str(dia), nome_cpf_tratado).mkdir(exist_ok=True)

    #localizar lupa
    pyautogui.moveTo(posicao_a_lupa)
    pyautogui.click()
    
    status = checar_conectividade()

    if status != "OK":
        print(f"⚠️ Problema de conexão: {status}")
        sys.exit()

    time.sleep(0.2)
    #Localizar anexo de documentos
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('./image/anexo_image.png', confidence = 0.8))
    x1, y1 = pyautogui.locateCenterOnScreen('./image/anexo_image.png', confidence = 0.8)
    x1 = x1 - 25

    #clicar nos documentos
    pyautogui.moveTo(x1, y1+27)
    for i in range(5):
        y1 = y1+23
        pyautogui.moveTo(x1, y1)
        time.sleep(0.5)
        pyautogui.click()
        time.sleep(0.2)
    time.sleep(0.2)
    pyautogui.click(pyautogui.locateCenterOnScreen('./image/X_image.png', confidence = 0.8))

    #Colocar documentos na pasta
    destino = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia),
    nome_cpf_tratado
)
    
    hoje = date.today()
    agora = datetime.now()

    #identificar erro no download
    for arquivo in downloads.iterdir():
        if arquivo.is_file():
            if arquivo.suffix != '.crdownload':
                data_arquivo = datetime.fromtimestamp(arquivo.stat().st_mtime)
                if agora - data_arquivo < timedelta(seconds=15):
                    time.sleep(0.2)
                    shutil.move(arquivo, destino / arquivo.name)

    #iteração das posições
    posicao_a_nome_S[1] = posicao_a_nome_S[1] + 40

    posicao_a_cpf_S[1] = posicao_a_cpf_S[1] + 40

    posicao_a_lupa[1] = posicao_a_lupa[1] + 40

#Abrir paginas no final da execução
caminho_pasta = Path(
    "Z:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia)
)

caminho_excel = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",str(ano), str(mes_nome), str(dia), "Controle de Solicitação.xlsx")

# os.startfile(caminho_csv)
os.startfile(caminho_pasta)
os.startfile(caminho_excel)