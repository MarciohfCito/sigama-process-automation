import pyautogui
import pyperclip
import time
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime

pyautogui.FAILSAFE = True

data_atual = datetime.today()


dia = data_atual.strftime("%d")
mes = data_atual.strftime("%m")
ano = data_atual.strftime("%Y")

print(dia, mes, ano)

#inserir o numero de registros
num = int(input("Digite o número de registros: "))

def minimizar():
    pyautogui.moveTo(pyautogui.moveTo(x=1806, y=7))
    pyautogui.click()
    time.sleep(0.5)

#minimizar vscode
minimizar()
time.sleep(0.5)

#POSIÇÕES
posicao_i_nome_S = [77, 336]
posicao_a_nome_S = posicao_i_nome_S

posicao_i_nome_E = [252, 309]
posicao_a_nome_E = posicao_i_nome_E

posicao_i_cpf_S = [839, 336]
posicao_a_cpf_S = posicao_i_cpf_S

posicao_i_cpf_E = [590, 309]
posicao_a_cpf_E = posicao_i_cpf_E
posicao_i_lupa = [1690, 336]
posicao_a_lupa = posicao_i_lupa

for j in range(num):

    #copiar nome - SIGAMA
    pyautogui.moveTo(posicao_a_nome_S)
    pyautogui.tripleClick()
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.5)

    arquivo = r"C:\SIGAMA\Documentos Solicitaçoes de Acesso\{ano}\{mes}\{dia}\Controle de Solicitação.xlsx"
    coluna = 'A'  # coluna desejada
    nome = pyperclip.paste()

    wb = load_workbook(arquivo)
    ws = wb["Controle de Solicitação"] # ou wb["NomeDaAba"]

    linha = 1

    # encontra a primeira célula vazia
    while ws[f"{coluna}{linha}"].value is not None:
        linha += 1

    # escreve o nome
    ws[f"{coluna}{linha}"] = nome

    # salva
    wb.save(arquivo)

    # #localizar excel
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\excel_image.png', confidence = 0.8))
    # time.sleep(0.5)

    # #colar nome EXCEL
    # pyautogui.moveTo(posicao_a_nome_E)
    # pyautogui.click()
    # time.sleep(0.5)
    # pyautogui.hotkey('ctrl', 'v')

    # #localizar chrome
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\chrome_image.png', confidence = 0.8))

    #copiar cpf SIGAMA
    pyautogui.moveTo(posicao_a_cpf_S)
    pyautogui.doubleClick()
    pyautogui.hotkey('ctrl', 'c')

    #cola cpf no excel
    coluna = 'B'
    cpf = pyperclip.paste()

    # escreve o cpf
    ws[f"{coluna}{linha}"] = cpf

    # salva
    wb.save(arquivo)

    colunas = ['A', 'B']   # colunas de destino

    # #localizar excel
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\excel_image.png', confidence = 0.8))

    # #colar cpf EXCEL
    # pyautogui.moveTo(posicao_a_cpf_E)
    # pyautogui.click()
    # time.sleep(0.2)
    # pyautogui.click()
    # pyautogui.hotkey('_')
    # time.sleep(0.2)
    # pyautogui.hotkey('ctrl', 'v')

    # # copiar nome e cpf EXCEL
    # pyautogui.moveTo(posicao_a_nome_E)
    # pyautogui.click()
    # time.sleep(0.5)
    # drag_x, drag_y = posicao_a_cpf_E
    # pyautogui.dragTo(drag_x, drag_y, 0.2, button='left')
    # pyautogui.hotkey('ctrl', 'c')
    # texto = pyperclip.paste()

    # texto_tratado = texto.replace("\t","").replace("\n", "").replace("\r", "")

    colunas = ["A", "B"]

    nome_cpf = [ws[f"{col}{linha}"].value for col in colunas]
    nome_cpf_tratado = nome_cpf.replace("\t", " - ")

    # #abrir gerenciador de arquivos e criar nova pasta
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\arquivos_image.png', confidence = 0.8))
    # time.sleep(0.5)
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\nova_pasta.png', confidence = 0.8))
    # time.sleep(0.5)
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\pasta.png', confidence = 0.8))
    # pyautogui.write(texto_tratado, interval=0.05)
    # pyautogui.hotkey('Enter')

    Path("C:\SIGAMA\Documentos Solicitaçoes de Acesso\{ano_atual}\{mes_atual}\{dia_atual}", nome_cpf_tratado).mkdir(exist_ok=True)

    # #localizar chrome
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\chrome_image.png', confidence = 0.8))

    #localizar lupa
    pyautogui.moveTo(posicao_a_lupa)
    pyautogui.click()
    time.sleep(1)

    #localizar documentos
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('.\image\anexo_image.png', confidence = 0.8))
    x1, y1 = pyautogui.locateCenterOnScreen('.\image\anexo_image.png', confidence = 0.8)

    #clicar no doc
    pyautogui.moveTo(x1, y1+27)

    for i in range(4):
        y1 = y1+23
        pyautogui.moveTo(x1, y1)
        time.sleep(0.5)
        pyautogui.click()
        time.sleep(0.2)

    pyautogui.click(pyautogui.locateCenterOnScreen('.\image\X_image.png', confidence = 0.8))

    # #abrir gerenciador de arquivos e downloads
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\arquivos_image.png', confidence = 0.8))
    # time.sleep(0.5)
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\download_image.png', confidence = 0.8))
    # time.sleep(0.5)
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\hoje_image.png', confidence = 0.8))
    # time.sleep(1)
    # pyautogui.hotkey('ctrl', 'c')
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\seta_voltar_image.png', confidence = 0.8))
    # time.sleep(0.5)
    # pyautogui.hotkey('Enter')
    # time.sleep(0.5)
    # pyautogui.hotkey('ctrl', 'v')
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\download_image.png', confidence = 0.8))
    # time.sleep(0.5)
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\hoje_image.png', confidence = 0.8))
    # time.sleep(1)
    # pyautogui.hotkey('Delete')
    # time.sleep(0.5)
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\seta_voltar_image.png', confidence = 0.8))
    # time.sleep(0.5)
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\seta_voltar_image.png', confidence = 0.8))
    # time.sleep(0.5)

    import shutil
    from datetime import date

    downloads = Path.home() / "Downloads"
    destino = Path("C:\SIGAMA\Documentos Solicitaçoes de Acesso\{ano_atual}\{mes_atual}\{dia_atual}", nome_cpf_tratado)

    destino.mkdir(exist_ok=True)

    hoje = date.today()

    for arquivo in downloads.iterdir():
        if arquivo.is_file():
            # data de modificação
            data_arquivo = date.fromtimestamp(arquivo.stat().st_mtime)

            if data_arquivo == hoje:
                shutil.move(arquivo, destino / arquivo.name)

    # #localizar chrome
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\chrome_image.png', confidence = 0.8))

    posicao_a_nome_S[1] = posicao_a_nome_S[1] + 45

    posicao_a_nome_E[1] = posicao_a_nome_E[1] + 26

    posicao_a_cpf_S[1] = posicao_a_cpf_S[1] + 45

    posicao_a_cpf_E[1] = posicao_a_cpf_E[1] + 26

    posicao_a_lupa[1] = posicao_a_lupa[1] + 45